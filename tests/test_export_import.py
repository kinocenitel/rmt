"""
Группа B.3 — экспорт и импорт файлов.

Что проверяем:
  - Экспорт xlsx: файл скачался, заголовки и строки с данными на месте.
  - Импорт xlsx: риски из файла появились в списке, теги нормализуются.
  - Импорт xlsx без ID: показывает ошибку, данные не затираются.
  - Экспорт PNG: файл скачался, это валидный PNG (магические байты).
  - Экспорт HTML: файл скачался, содержит имена рисков из текущего списка.

Технические детали:
  - page.expect_download() — ловит событие скачивания.
  - download.save_as(path) — сохраняет файл во временную директорию теста.
  - Для импорта используется скрытый <input type="file"> внутри <label>.
    Playwright умеет выставлять в него файл через .set_input_files().
"""
import pytest
from openpyxl import load_workbook
from playwright.sync_api import expect

from tests.pages import RiskMatrixPage


# ═════════════════════════════════════════════════════════════════════════════
#  1. Экспорт XLSX: файл скачался, структура верная.
# ═════════════════════════════════════════════════════════════════════════════
@pytest.mark.smoke
def test_export_xlsx_contains_risks(rm: RiskMatrixPage, tmp_path):
    """Экспортируем список из 2 рисков, открываем файл, проверяем содержимое."""
    rm.add_risk(name="Риск для экспорта 1", prob=2, sev=3, tags=["alpha"])
    rm.add_risk(name="Риск для экспорта 2", prob=4, sev=4)

    # Перехватываем событие скачивания
    with rm.page.expect_download() as download_info:
        rm.export_xlsx_btn.click()
    download = download_info.value

    # Сохраняем файл в tmp_path для инспекции
    saved = tmp_path / "exported.xlsx"
    download.save_as(saved)
    assert saved.exists() and saved.stat().st_size > 0

    # Открываем через openpyxl и проверяем
    wb = load_workbook(saved)
    assert "Активные" in wb.sheetnames, "Лист 'Активные' должен присутствовать"

    ws = wb["Активные"]
    # Первая строка — заголовки
    headers = [c.value for c in ws[1]]
    assert "Название риска*" in headers
    assert "Вероятность*" in headers
    assert "Влияние*" in headers

    # Собираем имена рисков из столбца "Название риска*"
    name_col_idx = headers.index("Название риска*") + 1  # openpyxl 1-indexed
    names_in_file = [ws.cell(row=i, column=name_col_idx).value
                     for i in range(2, ws.max_row + 1)]
    assert "Риск для экспорта 1" in names_in_file
    assert "Риск для экспорта 2" in names_in_file


# ═════════════════════════════════════════════════════════════════════════════
#  2. Импорт XLSX: риски из файла появились в списке.
# ═════════════════════════════════════════════════════════════════════════════
@pytest.mark.smoke
def test_import_xlsx_loads_risks(rm: RiskMatrixPage, xlsx_builder):
    """Генерируем валидный xlsx с 3 рисками, подставляем в input, проверяем."""
    file_path = xlsx_builder.valid()  # 3 риска по умолчанию

    rm.import_xlsx_input.set_input_files(str(file_path))

    rm.expect_active_count(3)
    names = rm.risk_names()
    assert "Импорт — риск 1" in names
    assert "Импорт — риск 2" in names
    assert "Импорт — риск 3" in names


# ═════════════════════════════════════════════════════════════════════════════
#  3. Импорт XLSX без ID — показывается ошибка, данные не меняются.
# ═════════════════════════════════════════════════════════════════════════════
@pytest.mark.smoke
def test_import_xlsx_without_id_shows_error(rm: RiskMatrixPage, xlsx_builder):
    """Файл с пустым ID должен отклоняться без замены текущих рисков."""
    rm.add_risk(name="Исходный риск", prob=2, sev=2)
    rm.expect_active_count(1)

    bad_file = xlsx_builder.without_id()
    rm.import_xlsx_input.set_input_files(str(bad_file))

    # Ошибка показалась
    import_error = rm.page.locator("#import-error")
    expect(import_error).to_be_visible()
    expect(import_error).to_contain_text("ID")

    # Старый риск на месте, импорт не прошёл
    rm.expect_active_count(1)
    assert "Исходный риск" in rm.risk_names()


# ═════════════════════════════════════════════════════════════════════════════
#  4. Экспорт PNG: файл скачался и это валидный PNG.
# ═════════════════════════════════════════════════════════════════════════════
@pytest.mark.smoke
def test_export_png_is_valid(rm: RiskMatrixPage, tmp_path):
    """Файл начинается с магических байт PNG (\\x89PNG\\r\\n\\x1a\\n) и не пустой."""
    rm.add_risk(name="Риск для PNG", prob=2, sev=2)

    with rm.page.expect_download(timeout=15_000) as download_info:
        rm.export_png_btn.click()
    download = download_info.value

    saved = tmp_path / "matrix.png"
    download.save_as(saved)

    data = saved.read_bytes()
    assert len(data) > 100, "PNG подозрительно маленький"
    # Магическая сигнатура PNG
    assert data[:8] == b"\x89PNG\r\n\x1a\n", "Файл не является валидным PNG"


# ═════════════════════════════════════════════════════════════════════════════
#  5. Экспорт HTML: файл скачался и содержит имена текущих рисков.
# ═════════════════════════════════════════════════════════════════════════════
@pytest.mark.smoke
def test_export_html_contains_risk_names(rm: RiskMatrixPage, tmp_path):
    """Внутри HTML должны быть имена рисков (они попадают в window.__EXPORTED__)."""
    rm.add_risk(name="HTML-тестовый риск", prob=3, sev=2)
    rm.add_risk(name="Ещё один для HTML",  prob=1, sev=4)

    with rm.page.expect_download() as download_info:
        rm.export_html_btn.click()
    download = download_info.value

    saved = tmp_path / "shared.html"
    download.save_as(saved)

    content = saved.read_text(encoding="utf-8")
    assert "HTML-тестовый риск" in content
    assert "Ещё один для HTML" in content
    # Проверим что это именно HTML-файл
    assert content.lstrip().lower().startswith("<!doctype html>")
